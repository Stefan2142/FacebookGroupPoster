# -*- coding: utf-8 -*-
import json,time, csv, os, random
from datetime import datetime
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import Select
from selenium.common.exceptions import NoSuchElementException
from selenium.common.exceptions import NoAlertPresentException
from selenium.common.exceptions import ElementNotVisibleException
from openpyxl import load_workbook

def output(name,message):
	with open(name+'.txt','w') as f:
		f.write(message.strip())

# Used for generating a list with full paths to images
def getFiles(path):
	fl = os.listdir(path)
	tmp = []
	for i in fl:
		tmp.append(path+i)
	# print tmp
	if tmp:
		return tmp
	else:
		return None
def main():
	# Check if file existst. If True, means that sript already run 
	# If False, this is the first run of the script and use (previousLine),
	# default value of '2' to start iterating through sheet. 
	# Because we are skiping the first row minus Header

	# PreviousLine is the number of the previous row scraped
	# From previous running of script

	username = ''
	password = ''
	line = '' # Current line
	previousLine = '2'
	fl = os.path.isfile("report.txt")
	if fl == True:
		with open('report.txt','r') as f:
			tmp = f.readline().strip()
			if tmp:
				previousLine = tmp
	elif fl == False:
		previousLine = '2' 

	# Load the excel workbook
	wb = load_workbook(filename = 'postit.xlsx')

	# Getting Data params
	data = wb['Data'] # 'Data' sheet
	imgLocation = data['A2'].value.strip().encode('utf8').replace('\\','\\\\')
	print 'Image location', imgLocation
	albumName = data['B2'].value.strip().encode('utf8')
	description = data['C2'].value.strip().encode('utf8')
	imgText = data['F2'].value.strip().encode('utf8')

	# Getting rows, urls for scraping
	links_range = wb['SourceLink']
	maxRow = links_range.max_row
	if previousLine == '2':
		rows = links_range['A'+str(int(previousLine)):'A'+str(maxRow)]
	else:
		rows = links_range['A'+str(int(previousLine)+1):'A'+str(maxRow)]
	# For Chrome webdriver
	chrome_options = webdriver.ChromeOptions()
	prefs = {"profile.default_content_setting_values.notifications" : 2}
	chrome_options.add_experimental_option("prefs",prefs)
	driver = webdriver.Chrome('C:\Users\stefa\Desktop\chromedriver.exe',chrome_options=chrome_options)

	driver.maximize_window()
	driver.get('https://www.facebook.com')
	driver.find_element_by_id('email').send_keys(username)
	driver.find_element_by_id('pass').send_keys(password)
	driver.find_element_by_id('loginbutton').click()
	time.sleep(1)


	# Main loop for iterating through rows
	for url in rows:
		try:
			currentRow = url.row
			driver.get(url.value)
			print 'Currently sraping:', url.value, currentRow
		except AttributeError:
			currentRow = url[0].row
			driver.get(url[0].value)
			print 'Currently scraping', url[0].value, currentRow
		
		driver.implicitly_wait(5)


		# Click on 'Add photo/Video button'
		# If button isnt found, you may be banned from group
		# Or group doesnt exist or some other reason. 
		# Exception is written to a provided xlsx
		try:
			driver.find_element_by_xpath("//*[@data-tooltip-content='Add Photo/Video']").click()
		except:
			try:
				driver.find_element_by_class_name('_44b4').click()
				time.sleep(1)
				driver.find_element_by_class_name('_5qtp').click()
			except Exception as e:
				print 'Cannot post to this group because %s'%(e)
				links_range.cell(row = currentRow, column = 2).value = str(datetime.now().strftime('%Y-%m-%d, %H:%M'))
				links_range.cell(row = currentRow, column = 3).value = 'Failure'
				links_range.cell(row= currentRow, column = 4).value = str(e)
				wb.save('postit-modified.xlsx')
				output('report',str(currentRow))
				delay = random.randint(5,12)
				print 'Url finished! Waiting for %s seconds.\n'%(str(delay))
				time.sleep(delay)
				continue


		time.sleep(5)
		driver.implicitly_wait(10)

		# Upload# Used for generating a list with full paths to imagesing images
		files = getFiles(imgLocation)
		files.sort(key=lambda f: int(filter(str.isdigit,f)))
		if files == None:
			output('error','No images found in the directory, %s'%(str(datetime.now().strftime('%Y-%m-%d %H %M'))))
			exit('No files loaded')
		for filePath in files:
			# print driver.find_elements_by_class_name('_3jk')
			try:
				uploader = driver.find_elements_by_class_name('_3jk')[1]
			except IndexError:
				uploader = driver.find_elements_by_class_name('_3jk')[0]
			inpt = uploader.find_element_by_xpath('.//input')
			inpt.send_keys(filePath)
			time.sleep(1)

		print 'Uploading images.'

		# Album title
		time.sleep(3)
		driver.find_element_by_xpath("//*[@placeholder='Album name']").send_keys(Keys.CONTROL+'a')
		driver.find_element_by_xpath("//*[@placeholder='Album name']").send_keys(Keys.DELETE)
		time.sleep(3)
		driver.find_element_by_xpath("//*[@placeholder='Album name']").send_keys(albumName)


		# Description
		print 'Updating description of the album'
		descr = driver.find_element_by_class_name('_4p8a')
		webdriver.ActionChains(driver).move_to_element(descr).click(descr).send_keys(description).perform()

		# Wait for 10 seconds before proceeding to add caption to each of images
		time.sleep(20) 

		# Image caption
		# caption = 'https://www.amazon.co.uk/dp/B01AS3KWM4'
		print 'Writing caption to each of the uploaded images.'
		captions = driver.find_elements_by_class_name('_5ipj')
		for i in captions:
			webdriver.ActionChains(driver).move_to_element(i).click(i).send_keys(imgText).perform()

		print 'Captions uploaded. Waiting for 10 seconds'
		time.sleep(5)
		driver.implicitly_wait(5)
		driver.find_element_by_xpath("//*[@data-testid='album-uploader-publish-button']").click() # Post 
		time.sleep(15)

		# Saving results
		links_range.cell(row = currentRow, column = 2).value = str(datetime.now().strftime('%Y-%m-%d, %H:%M'))
		links_range.cell(row = currentRow, column = 3).value = 'Success'
		wb.save('postit-modified.xlsx') #Overwrite postoji
		output('report',str(currentRow))
		delay = random.randint(50,1200)
		print 'Url finished! Waiting for %s seconds.\n'%(str(delay))
		time.sleep(delay)


if __name__ == '__main__':
	main()
